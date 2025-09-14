import os
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font

logger = logging.getLogger(__name__)

class Parser:
    """
    Parser for the record file
    """
    def __init__(self, quality_file_path: str, year_range: tuple):
        self.file_path = quality_file_path
        self.start_year = year_range[0]
        self.end_year = year_range[1]

    def parse_record(self):
        """
        Parse the record file
        """
        df = pd.read_csv(self.file_path)
        df = df.sort_values(by=['city', 'year', 'month'])
        city_list = df['city'].unique()

        total_year = range(self.start_year, self.end_year+1)
        cloud_high_color = Font(color='FF0000')
        cloud_median_color = Font(color='0000FF')
        cloud_low_color = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') # green for low cloud
        cover_high_color = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') # green for high cover
        cover_low_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # yellow for low cover
        note_city = Workbook()
        note_city.remove(note_city['Sheet'])
        for city in city_list:
            this_city_data = df[df['city'] == city]
            note_city_ws = note_city.create_sheet(title=city)
            current_row = 1
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
            logger.info('executing %s', city)
            with open(f"{self.file_path}/{city}.csv", "w", newline='', encoding='utf-8') as f:
                for year in total_year:
                    logger.info('executing %s', year)
                    f.write(','.join(map(str, self._date_line(year))) + '\n')
                    note_city_ws.append(self._date_line(year))
                    current_year_data = this_city_data[this_city_data['year'] == year]
                    property_list = ['toa_image_porpotion','sr_image_porpotion','toa_cloud_ratio','sr_cloud_ratio']
                    for pid, pro in enumerate(property_list):
                        value_list = ['/'] * 12
                        for _, row in current_year_data.iterrows():
                            value_list[row['month']-1] = row[pro]
                        row_line = [pro] + value_list
                        note_city_ws.append(row_line)
                        # change color of cloud ratio
                        if pro in ['toa_cloud_ratio', 'sr_cloud_ratio']:
                            for i in range(1, 13):
                                if row_line[i] != '/':
                                    if float(row_line[i]) > 10:
                                        note_city_ws.cell(column=i+1, row=current_row+pid+1).font = cloud_high_color
                                    elif float(row_line[i]) < 5:
                                        note_city_ws.cell(column=i+1, row=current_row+pid+1).fill = cloud_low_color
                                    else:
                                        note_city_ws.cell(column=i+1, row=current_row+pid+1).font = cloud_median_color
                        elif pro in ['toa_image_porpotion', 'sr_image_porpotion']:
                            for i in range(1, 13):
                                if row_line[i] != '/':
                                    if float(row_line[i]) < 0.9:
                                        note_city_ws.cell(column=i+1, row=current_row+pid+1).fill = cover_low_color
                                    else:
                                        note_city_ws.cell(column=i+1, row=current_row+pid+1).fill = cover_high_color

                        f.write(','.join(map(str, row_line)) + '\n')
                    current_row += 5
        note_city.save(os.path.join(self.file_path, 'city_quality_records.xlsx'))

    def _date_line(self, year):
        """
        Generate the date line
        """
        month = list(range(1, 13))
        return [year] + month
